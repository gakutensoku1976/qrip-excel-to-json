import json
import openpyxl
import os
import sys
from datetime import datetime
import warnings

INPUT_SHEET_NAME = '入力欄'
JSON_VERSION = '20230910'

def is_running_from_cmd():
    try:
        # コマンドプロンプトから実行されている場合、sys.stdin は接続されています
        return sys.stdin.isatty()
    except AttributeError:
        # sys.stdin.isatty() が利用できない場合は、ダブルクリックなどで直接実行されたとみなします
        return False

def ConvertExcelDir(directory):
    xlsx_files = [file for file in os.listdir(directory) if file.endswith('.xlsx')]

    if len(xlsx_files) > 0:

        exe_flag = True

        # コマンドプロンプトから実行されている場合
        if is_running_from_cmd():
            print('処理対象ファイル')
            for xlsx_file in xlsx_files:
                print(xlsx_file)

            user_input = input("\n変換処理を実行しますか?(Y/n): ")

            if user_input.lower() == 'y':
                print("変換処理を実行します。\n")
                exe_flag = True
            elif user_input.lower() == 'n':
                print("変換処理を実行せずにプログラムを終了します。")
                exe_flag = False
            else:
                print(f"不正な入力({user_input})です。プログラムを終了します。")
                exe_flag = False

        if exe_flag:
            for xlsx_file in xlsx_files:
                Excel2Json(os.path.join(directory, xlsx_file))

        return 0
    else:
        print(f"Error: カレントフォルダにExcelファイルがありません")
        return 0

def print_args(argv):
    for option in argv:
        print(option)

def IsQripExcelFormat(active_sheet):

    if active_sheet.title == '入力欄' or str(active_sheet['A1'].value) == JSON_VERSION:
        # 参加者が二人以上いるか?
        if active_sheet['A19'].value is None or active_sheet['A20'].value is None:
            return False
        else:
            return True
    else:
        return False

# 名前文字列の正規化 (全角半角スペースを削除、全角英数字を半角に変換、全角記号を半角に変換)
def normalize_name(name):
    name = name.replace(' ', '')
    name = name.translate(str.maketrans('０１２３４５６７８９','0123456789'))
    name = name.translate(str.maketrans('ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ','abcdefghijklmnopqrstuvwxyz'))
    name = name.translate(str.maketrans('ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ','ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
    name = name.translate(str.maketrans('！＂＃＄％＆＇（）＊＋，－．／：；＜＝＞？＠［￥］＾＿｀｛｜｝～','!"#$%&\'()*+,-./:;<=>?@[¥]^_`{|}~'))
    return name

def Sheet2Json(active_sheet,excel_file_path):

    qric_json = {}

    # 'title'
    qric_json[active_sheet['B2'].value] = active_sheet['C2'].value
    qric_json['version'] = JSON_VERSION

    # メタデータ
    METADATA_START_ROW = 3
    METADATA_END_ROW = 16

    meta_data = {}
    for row_num in range(METADATA_START_ROW,METADATA_END_ROW):

        name = active_sheet.cell(row=row_num, column=2).value

        value = active_sheet.cell(row=row_num, column=3).value

        if name == 'period':
            # 'period' は 'end' - 'start' で計算される値であるため、変換時には無視して None を設定する
            meta_data[name] = None
        elif isinstance(value, datetime): 
            meta_data[name] = datetime.strftime(value, "%Y/%m/%d %H:%M")
        else:
            meta_data[name] = value

    qric_json['meta'] = meta_data

    player_list = []
    row_number = 19
    while active_sheet.cell(row=row_number, column=1).value is not None:

        places_data = {}

        places_data['name'] = normalize_name(str(active_sheet.cell(row=row_number, column=1).value))
        
        result_data = {
           'rank': active_sheet.cell(row=row_number, column=2).value,
           'round': active_sheet.cell(row=row_number, column=3).value,
           'point': active_sheet.cell(row=row_number, column=4).value,
           'maru': active_sheet.cell(row=row_number, column=5).value,
           'miss': active_sheet.cell(row=row_number, column=6).value,
           'winout': active_sheet.cell(row=row_number, column=7).value,
           'fail': active_sheet.cell(row=row_number, column=8).value,
           'comment': active_sheet.cell(row=row_number, column=9).value
           }

        places_data['result'] = result_data

        player_list.append(places_data)
        row_number += 1

    match_data = {}
    match_data['places'] = player_list
    qric_json['match'] = match_data

    parts = excel_file_path.split('.')
    parts[-1] = active_sheet.title + '.json'
    json_file_path = '_'.join(parts)

    try:
        with open(json_file_path, 'w', encoding='utf-8') as file:
            json.dump(qric_json, file, indent=4, ensure_ascii=False)
            if is_running_from_cmd():
                print(f'【出力】{json_file_path}')
    except Exception as e:
        print(f"ファイル書き込み時にエラー発生: {e}")

def Excel2Json(excel_file_path):
    try:
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    except Exception as e:
        print(f"予期しないエラーが発生しました: {e}")

    for sheet_name in workbook.sheetnames:
        active_sheet = workbook[sheet_name]            
        if IsQripExcelFormat(active_sheet):
            Sheet2Json(active_sheet,excel_file_path)

    workbook.close()

def main():
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

    if len(sys.argv) == 1:
        # 引数が指定されていなければ、カレントディレクトリに対して処理
        ConvertExcelDir(os.getcwd())

    elif len(sys.argv) == 2:
        # 引数が1個だったら、Excelファイルか対象フォルダ
        target_path = sys.argv[1]
        if os.path.isfile(target_path):
            parts = target_path.split('.')
            if parts[-1] == 'xlsx':
                Excel2Json(target_path)
                return 0
            else :
                print(f"Error: {target_path} はExcelファイルではありません")
                return -1
        elif os.path.isdir(target_path):
            ConvertExcelDir(target_path)
            return 0
        else:
            print(f"Error: パラメータの指定が不正です {target_path} ")
            return -1
    else:
        print(f"Error: パラメータの指定が不正です {sys.argv} ")

if __name__ == "__main__":
    main()